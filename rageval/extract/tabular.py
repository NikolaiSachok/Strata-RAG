"""Spreadsheet reading (#41) — pluggable xlsx/csv readers that yield STRUCTURED, provenance-
carrying rows. Corpus-neutral: it parses rows/cells and knows NO column names.

WHY spreadsheets are handled COMPLETELY differently from documents (the load-bearing decision):
  A policy PDF is narrative → you EMBED it and retrieve passages. A loss-run spreadsheet is
  STRUCTURED, AGGREGATABLE data → the useful questions are "how many open claims?", "total reserve
  by peril", "which agent wrote the most policies?". A vector top-k CANNOT count, sum, or group by;
  and embedding table rows as prose DILUTES retrieval (dozens of near-identical "row" chunks crowd
  out the real narrative). So spreadsheet rows do NOT become embedded chunks. They become
  STRUCTURED FACTS in the sidecar (facts.py / the EAV store), where the existing `aggregate` path
  answers count/list/group_by/lookup deterministically. This is the same "structured data → SQL
  facet, not vector search" split the config.yaml harvest already uses — now for tabular files.

THE CORE / ADAPTER BOUNDARY (why no column name lives here):
  This module is the generic MECHANISM: open the file, find the header row, yield each data row as
  an ORDERED {column_name: cell_value} mapping WITH its sheet + 1-based row number. It has no
  opinion about what "Premium" or "Cause of loss" means — mapping a spreadsheet's columns to typed
  facets is the corpus adapter's business (`declared_facets()` / `harvest_facts()`), exactly like
  the descriptor whitelist. So the engine core carries zero corpus-specific column vocabulary
  (grep-provable), and a new corpus with a totally different sheet supplies its own mapping with no
  core edit.

PLUGGABILITY: xlsx via openpyxl (lazy-imported, read-only mode), csv via the stdlib. Both produce
the SAME neutral `TabularData` shape, so the adapter's harvester is reader-agnostic and a third
format (ods, parquet) is a new `TabularReader` subclass with nothing downstream changing.
"""

from __future__ import annotations

import abc
import csv
import io
from dataclasses import dataclass, field
from pathlib import Path


@dataclass(frozen=True)
class TabularRow:
    """One data row, normalised to a neutral shape carrying its own provenance.

    Fields:
      sheet   — the sheet/tab name ("" or the file stem for single-sheet csv).
      index   — 1-based row number WITHIN the data rows (header excluded) — stable row provenance.
      cells   — ORDERED {header: value} for this row. Header names are taken verbatim from the
                file's header row; values are strings/None (numeric coercion is the ADAPTER's job,
                against its declared facet TYPES — the reader stays type-agnostic).
    """

    sheet: str
    index: int
    cells: dict[str, object]

    def get(self, column: str, default: object = None) -> object:
        return self.cells.get(column, default)


@dataclass(frozen=True)
class TabularData:
    """A parsed spreadsheet: its headers + data rows, plus which backend read it (auditability).

    A single flat list of rows across all sheets (each row carries its `sheet`), so an adapter can
    iterate rows uniformly and still know the sheet provenance."""

    headers: tuple[str, ...]
    rows: tuple[TabularRow, ...]
    backend: str = ""
    meta: dict = field(default_factory=dict)

    @property
    def row_count(self) -> int:
        return len(self.rows)

    @property
    def sheets(self) -> tuple[str, ...]:
        seen: list[str] = []
        for r in self.rows:
            if r.sheet not in seen:
                seen.append(r.sheet)
        return tuple(seen)


# --- resource caps (MAJOR-3, DoS): bound what a single file can materialise -------------------
# An untrusted spreadsheet can be enormous (200k rows / hundreds of MB). Streaming avoids building
# the whole sheet in memory, and these caps bound the WORK regardless: a file beyond them is
# TRUNCATED (its extra rows skipped) with the truncation SURFACED in TabularData.meta so a caller
# can flag it (like the scanned-PDF coverage warning) — never a silent OOM. Corpus-neutral defaults;
# an adapter/caller can override per read.
@dataclass(frozen=True)
class TabularLimits:
    """Caps for one spreadsheet read (MAJOR-3). max_rows/max_cols bound the parsed shape;
    max_bytes bounds the raw csv slurp. Reaching a cap TRUNCATES + records it in meta['truncated']
    rather than raising, so a huge file degrades gracefully instead of OOMing."""

    max_rows: int = 100_000       # data rows per file (across sheets)
    max_cols: int = 512           # columns considered per row
    max_bytes: int = 64 * 1024 * 1024  # 64 MiB cap on a csv read


DEFAULT_TABULAR_LIMITS = TabularLimits()


# The banner-vs-header decision keys off SHAPE, never corpus words. A leading row is skipped as a
# BANNER only when a LATER row is STRICTLY WIDER — i.e. the leading row is narrower than the real
# table (a 1-cell "SYNTHETIC — …" line above a multi-column header). A genuinely single-COLUMN sheet
# (claim ids, policy numbers) has no wider row below, so its first non-empty row is the header and
# its single column is KEPT (CRITICAL-3: never silently drop a 1-column sheet).
def _row_width(cells: list) -> int:
    """Number of non-empty cells in a row (its effective width)."""
    return sum(1 for c in cells if c is not None and str(c).strip())


def _dedupe_headers(raw: list[str]) -> list[str]:
    """Normalise a header row to non-empty, unique column names (stable, deterministic).

    Blank headers become `column_<n>`; a duplicate name gets a `_<n>` suffix. This guarantees the
    per-row {header: value} mapping has no colliding/empty keys, so the adapter can address every
    column unambiguously."""
    out: list[str] = []
    seen: dict[str, int] = {}
    for i, name in enumerate(raw):
        base = str(name).strip() if name is not None else ""
        if not base:
            base = f"column_{i + 1}"
        if base in seen:
            seen[base] += 1
            base = f"{base}_{seen[base]}"
        else:
            seen[base] = 0
        out.append(base)
    return out


def _pick_header(head_rows: list[list]) -> int | None:
    """Choose the header-row index from the first few non-blank rows (STREAMING-friendly: only a
    small lookahead window is needed).

    Rule (shape-only, CRITICAL-3): the header is the FIRST non-blank row, UNLESS it is a narrower
    leading BANNER — i.e. a later row in the window is STRICTLY WIDER — in which case the first row
    at that wider width is the header. So a 1-cell "SYNTHETIC" banner above a 5-column table is
    skipped, but a genuinely single-column sheet (no wider row) keeps its first row as the header."""
    widths = [(_row_width(r)) for r in head_rows]
    non_blank = [i for i, w in enumerate(widths) if w > 0]
    if not non_blank:
        return None
    first = non_blank[0]
    max_w = max(widths)
    if widths[first] < max_w:
        # The first non-blank row is narrower than a later row → it's a banner; the header is the
        # first row that reaches the max width.
        for i in non_blank:
            if widths[i] == max_w:
                return i
    return first


def _make_row(sheet: str, headers: list[str], raw: list, data_index: int) -> TabularRow:
    cells: dict[str, object] = {}
    for col_i, header in enumerate(headers):
        value = raw[col_i] if col_i < len(raw) else None
        if isinstance(value, str):
            value = value.strip() or None
        cells[header] = value
    return TabularRow(sheet=sheet, index=data_index, cells=cells)


def _rows_from_iter(sheet: str, row_iter, *, limits: TabularLimits,
                    rows_budget: int) -> tuple[list[str], list[TabularRow], bool]:
    """Stream `row_iter` (an iterable of cell-lists) → (headers, TabularRows, truncated).

    Header detection uses a SMALL lookahead window (so we never materialise the whole sheet), then
    each subsequent non-blank row is emitted lazily. Stops after `rows_budget` data rows (the
    remaining per-file cap) and reports `truncated=True` if more rows were available. Cells beyond
    max_cols are ignored. A merged-cell / short row pads missing columns with None."""
    it = iter(row_iter)
    window: list[list] = []
    truncated = False
    # Gather a small lookahead of the first non-blank-ish rows to decide the header.
    LOOKAHEAD = 8
    for raw in it:
        raw = list(raw)[: limits.max_cols]
        window.append(raw)
        if len([r for r in window if _row_width(r) > 0]) >= 2 or len(window) >= LOOKAHEAD:
            break
    header_idx = _pick_header(window)
    if header_idx is None:
        # No non-blank row in the window → maybe everything so far was blank; try the rest lazily.
        for raw in it:
            raw = list(raw)[: limits.max_cols]
            if _row_width(raw) > 0:
                window.append(raw)
                header_idx = len(window) - 1
                break
        if header_idx is None:
            return [], [], False
    headers = _dedupe_headers([("" if c is None else str(c)) for c in window[header_idx]])
    rows: list[TabularRow] = []
    data_index = 0

    def _emit(raw: list) -> bool:
        """Emit one data row; return False when the budget is exhausted."""
        nonlocal data_index, truncated
        if not any(c is not None and str(c).strip() for c in raw):
            return True  # skip a fully-blank row (trailing filler)
        if data_index >= rows_budget:
            truncated = True
            return False
        data_index += 1
        rows.append(_make_row(sheet, headers, raw, data_index))
        return True

    # The rows already pulled into the window AFTER the header.
    for raw in window[header_idx + 1:]:
        if not _emit(raw):
            return headers, rows, truncated
    # Then the streamed remainder.
    for raw in it:
        raw = list(raw)[: limits.max_cols]
        if not _emit(raw):
            return headers, rows, truncated
    return headers, rows, truncated


class TabularReader(abc.ABC):
    """The swappable spreadsheet-reading interface. A backend implements `read`."""

    name: str = "base"

    @abc.abstractmethod
    def read(self, path: Path, *, limits: TabularLimits = DEFAULT_TABULAR_LIMITS) -> TabularData:
        raise NotImplementedError


class XlsxReader(TabularReader):
    """xlsx via openpyxl in READ-ONLY, STREAMING mode (MAJOR-3: never materialises a whole sheet).

    read_only=True streams rows without loading the whole workbook; data_only=True returns the last
    cached cell VALUES rather than formulae (we want data, not `=SUM(...)`). Every sheet is read;
    each row keeps its sheet name as provenance. A per-FILE row budget is shared across sheets so a
    huge workbook is bounded, with any truncation recorded in meta."""

    name = "openpyxl"

    def read(self, path: Path, *, limits: TabularLimits = DEFAULT_TABULAR_LIMITS) -> TabularData:
        from openpyxl import load_workbook  # lazy: only paid when an xlsx is actually read

        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        all_headers: list[str] = []
        all_rows: list[TabularRow] = []
        truncated = False
        try:
            for ws in wb.worksheets:
                budget = limits.max_rows - len(all_rows)
                if budget <= 0:
                    truncated = True
                    break
                # ws.iter_rows(values_only=True) is a LAZY generator in read-only mode → streamed.
                headers, rows, tr = _rows_from_iter(
                    ws.title, ws.iter_rows(values_only=True), limits=limits, rows_budget=budget)
                truncated = truncated or tr
                if headers and not all_headers:
                    all_headers = headers
                all_rows.extend(rows)
        finally:
            wb.close()  # read-only workbooks hold a file handle until closed
        return TabularData(headers=tuple(all_headers), rows=tuple(all_rows), backend=self.name,
                           meta={"sheet_count": len({r.sheet for r in all_rows}),
                                 "truncated": truncated})


class CsvReader(TabularReader):
    """csv via the stdlib. Single logical sheet (named after the file stem). Sniffs the delimiter
    with a tolerant fallback to comma; decodes UTF-8 leniently. Reads at most max_bytes and streams
    rows through the parser (MAJOR-3: never slurps an unbounded file into a full matrix)."""

    name = "csv"

    def read(self, path: Path, *, limits: TabularLimits = DEFAULT_TABULAR_LIMITS) -> TabularData:
        # Bounded read: pull at most max_bytes + 1 so we can DETECT (and flag) an over-cap file
        # without loading the rest of it into memory.
        with open(path, "r", encoding="utf-8", errors="replace") as fh:
            raw = fh.read(limits.max_bytes + 1)
        byte_truncated = len(raw) > limits.max_bytes
        if byte_truncated:
            raw = raw[: limits.max_bytes]
        # Sniff the dialect from a sample; fall back to a plain comma dialect if sniffing fails
        # (a one-column banner line can defeat the sniffer — comma is the safe default).
        try:
            dialect: type[csv.Dialect] | csv.Dialect = csv.Sniffer().sniff(raw[:4096], delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        sheet = Path(path).stem
        headers, rows, row_truncated = _rows_from_iter(
            sheet, csv.reader(io.StringIO(raw), dialect), limits=limits, rows_budget=limits.max_rows)
        return TabularData(headers=tuple(headers), rows=tuple(rows), backend=self.name,
                           meta={"sheet_count": 1, "truncated": byte_truncated or row_truncated})


# Extension → reader. The generic dispatch point; a new tabular format registers here.
_READERS: dict[str, TabularReader] = {
    "xlsx": XlsxReader(),
    "csv": CsvReader(),
}


def read_tabular(path: Path, *, reader: TabularReader | None = None,
                 limits: TabularLimits = DEFAULT_TABULAR_LIMITS) -> TabularData:
    """Read a spreadsheet by extension (xlsx/csv), or with an explicit `reader`. Raises ValueError
    for an unsupported extension so a mis-wired adapter fails loud rather than silently empty.
    `limits` bounds rows/cols/bytes (MAJOR-3); truncation is recorded in the result's meta."""
    path = Path(path)
    if reader is not None:
        return reader.read(path, limits=limits)
    ext = path.suffix.lower().lstrip(".")
    chosen = _READERS.get(ext)
    if chosen is None:
        raise ValueError(f"no tabular reader for .{ext} (supported: {sorted(_READERS)})")
    return chosen.read(path, limits=limits)
